#TGVMAXPLUS_simple_final_1.0_stable

#---------------




#import de toutes les librairies nécessaires
import requests
import json
import smtplib
import xlrd
from email.mime.text import MIMEText
import xlwt
from openpyxl import load_workbook
import time



#Recherche de TGVMAX sur la base de donnée SNCF
def recherche(L): 
            
    #requete à l'API SNCF
    
    r=requests.get('https://data.sncf.com/api/records/1.0/search/?dataset=tgvmax&sort=date&facet=date&facet=origine&facet=destination&facet=heure_depart&refine.date={}%2F{}%2F{}&refine.origine={}&refine.destination={}&refine.heure_depart={}%3A{}'.format(L[7],L[6],L[5],L[1],L[2],L[3],L[4]))
    
    #extraction du nombre de TGVMax disponibles (1 ou O)
    x=r.json()['nhits']
    
    return x     




def avertissement(L,k):
        #modification dans l'excel de commande de l'état de la commande
        
        wb = load_workbook('C:\\Users\\leand\\Desktop\\Documents\\Python Scripts\\TGVMAXPLUS\\TGVMAXPLUS.xlsx')
        ws = wb.active
        ws['I{}'.format(k+1)] = 1
        wb.save("C:\\Users\\leand\\Desktop\\Documents\\Python Scripts\\TGVMAXPLUS\\TGVMAXPLUS.xlsx")
        
        #envoi d'un email d'information (délai de 4 min)
        server = smtplib.SMTP('smtp.mail.yahoo.com', 587)
        server.starttls()
        server.login("tgvmaxplus@yahoo.com", "Leandre97")
        msg = MIMEText("TGV Max de {} à {} pour {}:{} le {}/{}/{} libéré".format(L[1],L[2],L[3],L[4],L[5],L[6],L[7]))
        msg['Subject']= 'TGVMAX libéré !'
        msg['From']='tgvmaxplus@yahoo.com'
        msg['To']='{}'.format(L[12])
        server.sendmail("tgvmaxplus@yahoo.com", "{}".format(L[12]), msg.as_string())
        server.quit()
        
        return
    
    
def main():    
    #ouverture de l'excel de commande   
    data=xlrd.open_workbook('C:\\Users\\leand\\Desktop\\Documents\\Python Scripts\\TGVMAXPLUS\\TGVMAXPLUS.xlsx')
    
    #ouverture de la feuille active
    requetes=data.sheet_by_name(u'Requetes')
    
    
    #boucle sur les lignes de l'excel de commande
    for k in range(1,requetes.nrows) :
        
        #création d'une liste et d'un dictionnaire vide pour recevoir respectivement les commabdes et le résultat de la requète à l'API et les coordonées
        L=[]
        r={}
        
        #extraction des commandes
        L=requetes.row_values(k)
        
        #extraction de la commande de réservation
        resa = L[9]
                
        #formatage des valeurs des heures et minutes et date
                
        for w in range(3,8):
            L[w]=int(L[w])
            print(L[w])
                
        
            #vérification de la validité de la commande : 0 si non traitée, 1 sinon
            
        if L[8]==0:
            
            x = recherche(L)
            if x==1:
                avertissement(L,k)
            else:
                return 0
        else:
            return 0
            
    return 1


    

